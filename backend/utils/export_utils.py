"""
Export Utilities for Attendance System
Handles Excel, CSV, and PDF export functionality.
"""
import os
import io
import csv
import logging
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from fastapi import HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from database import Student, Class, AttendanceSession, AttendanceRecord

logger = logging.getLogger(__name__)


class AttendanceExporter:
    """Handles all attendance export functionality"""
    
    def __init__(self):
        self.exports_dir = "/home/bitbuggy/Naman_Projects/Dental Attendance/backend/static/exports"
        os.makedirs(self.exports_dir, exist_ok=True)
    
    def get_date_filters(self, period: str) -> tuple[datetime, datetime, str]:
        """Calculate date filters based on period"""
        end_date = datetime.now()
        if period == "weekly":
            start_date = end_date - timedelta(weeks=1)
            period_name = "Weekly"
        elif period == "monthly":
            start_date = end_date - timedelta(days=30)
            period_name = "Monthly"
        else:  # all
            start_date = datetime(2020, 1, 1)
            period_name = "Complete"
        return start_date, end_date, period_name
    
    def get_class_name(self, class_id: Optional[int], db: Session) -> str:
        """Get formatted class name for filename"""
        if class_id:
            class_obj = db.query(Class).filter(Class.id == class_id).first()
            if class_obj:
                return f"{class_obj.name}_{class_obj.section}".replace(" ", "_")
        return "All_Classes"
    
    def get_filtered_sessions(self, start_date: datetime, end_date: datetime, 
                            class_id: Optional[int], db: Session) -> List[AttendanceSession]:
        """Get filtered attendance sessions"""
        session_query = db.query(AttendanceSession).filter(
            AttendanceSession.created_at >= start_date,
            AttendanceSession.created_at <= end_date
        )
        if class_id:
            session_query = session_query.filter(AttendanceSession.class_id == class_id)
        return session_query.order_by(AttendanceSession.created_at.desc()).all()
    
    def get_student_analytics(self, start_date: datetime, end_date: datetime,
                            class_id: Optional[int], db: Session) -> List[Dict[str, Any]]:
        """Get student attendance analytics"""
        students_query = db.query(Student).filter(Student.is_active == True)
        if class_id:
            students_query = students_query.filter(Student.class_id == class_id)
        
        student_analytics = []
        for student in students_query.all():
            # Get student's attendance in the period
            student_records = db.query(AttendanceRecord).join(AttendanceSession).filter(
                AttendanceRecord.student_id == student.id,
                AttendanceSession.created_at >= start_date,
                AttendanceSession.created_at <= end_date
            )
            if class_id:
                student_records = student_records.filter(AttendanceSession.class_id == class_id)
            
            total_sessions = student_records.count()
            present_sessions = student_records.filter(AttendanceRecord.is_present == True).count()
            
            if total_sessions > 0:
                attendance_rate = round((present_sessions / total_sessions) * 100, 1)
                status = "🟢 Excellent" if attendance_rate >= 90 else "🟡 Good" if attendance_rate >= 75 else "🟠 Average" if attendance_rate >= 60 else "🔴 Poor"
                
                student_analytics.append({
                    'Student Name': student.name,
                    'Roll No': student.roll_no,
                    'Total Sessions': total_sessions,
                    'Present': present_sessions,
                    'Absent': total_sessions - present_sessions,
                    'Attendance %': attendance_rate,
                    'Status': status
                })
        
        # Sort by attendance rate
        student_analytics.sort(key=lambda x: x['Attendance %'], reverse=True)
        return student_analytics
    
    def create_excel_styles(self):
        """Create Excel styling objects"""
        return {
            'header_font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),
            'subheader_font': Font(name='Calibri', size=11, bold=True, color='2E75B6'),
            'normal_font': Font(name='Calibri', size=10),
            'center_align': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def create_summary_sheet(self, wb: Workbook, sessions: List[AttendanceSession], 
                           start_date: datetime, end_date: datetime, period_name: str,
                           class_id: Optional[int], db: Session, styles: Dict):
        """Create the attendance summary sheet"""
        ws_summary = wb.create_sheet("📊 Attendance Summary")
        
        # Header section
        ws_summary['A1'] = f"🎓 Attendance Report - {period_name}"
        ws_summary['A1'].font = Font(name='Calibri', size=16, bold=True, color='2E75B6')
        ws_summary.merge_cells('A1:F1')
        
        ws_summary['A2'] = f"📅 Period: {start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
        ws_summary['A2'].font = Font(name='Calibri', size=11, color='666666')
        ws_summary.merge_cells('A2:F2')
        
        if class_id:
            class_obj = db.query(Class).filter(Class.id == class_id).first()
            ws_summary['A3'] = f"🏫 Class: {class_obj.name} - Section {class_obj.section}"
            ws_summary['A3'].font = Font(name='Calibri', size=11, color='666666')
            ws_summary.merge_cells('A3:F3')
            start_row = 5
        else:
            start_row = 4
        
        # Summary headers
        headers = ['📅 Date', '⏰ Session', '🏫 Class', '👥 Total', '✅ Present', '📊 Rate %']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=start_row, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['center_align']
            cell.border = styles['border']
        
        # Summary data
        for idx, session in enumerate(sessions, start_row + 1):
            class_obj = db.query(Class).filter(Class.id == session.class_id).first()
            class_display = f"{class_obj.name} {class_obj.section}" if class_obj else "Unknown"
            
            attendance_rate = round((session.total_present / max(1, session.total_detected)) * 100, 1)
            
            row_data = [
                session.created_at.strftime('%d-%m-%Y'),
                session.session_name[:30] + '...' if len(session.session_name) > 30 else session.session_name,
                class_display,
                session.total_detected,
                session.total_present,
                f"{attendance_rate}%"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=idx, column=col, value=value)
                cell.font = styles['normal_font']
                cell.alignment = styles['center_align']
                cell.border = styles['border']
                
                # Color coding for attendance rate
                if col == 6:  # Attendance rate column
                    if attendance_rate >= 80:
                        cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    elif attendance_rate >= 60:
                        cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        
        # Auto-fit columns
        self.auto_fit_columns(ws_summary)
    
    def create_analytics_sheet(self, wb: Workbook, student_analytics: List[Dict], styles: Dict):
        """Create the student analytics sheet"""
        ws_analytics = wb.create_sheet("📈 Student Analytics")
        
        # Header
        ws_analytics['A1'] = f"📈 Student Performance Analytics"
        ws_analytics['A1'].font = Font(name='Calibri', size=16, bold=True, color='2E75B6')
        ws_analytics.merge_cells('A1:G1')
        
        # Analytics headers
        analytics_headers = ['👤 Student Name', '🆔 Roll No', '📅 Total Sessions', '✅ Present', '❌ Absent', '📊 Attendance %', '⭐ Status']
        for col, header in enumerate(analytics_headers, 1):
            cell = ws_analytics.cell(row=3, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['center_align']
            cell.border = styles['border']
        
        # Analytics data
        for idx, student_data in enumerate(student_analytics, 4):
            for col, (key, value) in enumerate(student_data.items(), 1):
                cell = ws_analytics.cell(row=idx, column=col, value=value)
                cell.font = styles['normal_font']
                cell.alignment = styles['center_align']
                cell.border = styles['border']
                
                # Color coding based on attendance percentage
                if key == 'Attendance %':
                    if value >= 90:
                        cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    elif value >= 75:
                        cell.fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')
                    elif value >= 60:
                        cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        
        # Auto-fit columns
        self.auto_fit_columns(ws_analytics, max_width=30)
    
    def create_detailed_sheet(self, wb: Workbook, sessions: List[AttendanceSession], 
                            db: Session, styles: Dict):
        """Create the detailed records sheet"""
        ws_detailed = wb.create_sheet("📋 Detailed Records")
        
        # Get all attendance records
        detailed_records = []
        for session in sessions:
            records = db.query(AttendanceRecord).join(Student).filter(
                AttendanceRecord.session_id == session.id
            ).all()
            
            class_obj = db.query(Class).filter(Class.id == session.class_id).first()
            class_name = f"{class_obj.name} {class_obj.section}" if class_obj else "Unknown"
            
            for record in records:
                detailed_records.append({
                    'Date': session.created_at.strftime('%d-%m-%Y'),
                    'Session': session.session_name,
                    'Class': class_name,
                    'Student': record.student.name,
                    'Roll No': record.student.roll_no,
                    'Status': '✅ Present' if record.is_present else '❌ Absent',
                    'Confidence': f"{record.confidence:.2f}" if record.confidence else "N/A"
                })
        
        # Headers for detailed records
        ws_detailed['A1'] = "📋 Detailed Attendance Records"
        ws_detailed['A1'].font = Font(name='Calibri', size=16, bold=True, color='2E75B6')
        ws_detailed.merge_cells('A1:G1')
        
        detailed_headers = ['📅 Date', '📝 Session', '🏫 Class', '👤 Student', '🆔 Roll No', '✅ Status', '🎯 Confidence']
        for col, header in enumerate(detailed_headers, 1):
            cell = ws_detailed.cell(row=3, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['center_align']
            cell.border = styles['border']
        
        # Detailed data
        for idx, record in enumerate(detailed_records, 4):
            for col, (key, value) in enumerate(record.items(), 1):
                cell = ws_detailed.cell(row=idx, column=col, value=value)
                cell.font = styles['normal_font']
                cell.alignment = styles['center_align']
                cell.border = styles['border']
        
        # Auto-fit columns
        self.auto_fit_columns(ws_detailed)
    
    def auto_fit_columns(self, worksheet, max_width: int = 25):
        """Auto-fit columns to content with maximum width limit"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    async def export_excel(self, period: str, class_id: Optional[int], 
                          format_type: str, db: Session) -> FileResponse:
        """Export beautiful, organized attendance data to Excel"""
        try:
            # Get date filters and class name
            start_date, end_date, period_name = self.get_date_filters(period)
            class_name = self.get_class_name(class_id, db)
            
            # Get filtered data
            sessions = self.get_filtered_sessions(start_date, end_date, class_id, db)
            if not sessions:
                raise HTTPException(status_code=404, detail="No attendance data found for the specified criteria")
            
            # Create workbook and styles
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            styles = self.create_excel_styles()
            
            # Create sheets
            self.create_summary_sheet(wb, sessions, start_date, end_date, period_name, class_id, db, styles)
            
            # Get student analytics
            student_analytics = self.get_student_analytics(start_date, end_date, class_id, db)
            self.create_analytics_sheet(wb, student_analytics, styles)
            
            # Create detailed sheet if requested
            if format_type == "detailed":
                self.create_detailed_sheet(wb, sessions, db, styles)
            
            # Save the workbook
            filename = f"Attendance_{period_name}_{class_name}_{end_date.strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.exports_dir, filename)
            wb.save(filepath)
            
            logger.info(f"Beautiful Excel export created: {filename}")
            
            return FileResponse(
                path=filepath,
                filename=filename,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Excel export error: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Failed to export attendance data: {str(e)}")
    
    async def export_csv(self, period: str, class_id: Optional[int], db: Session) -> StreamingResponse:
        """Export attendance data to CSV format"""
        try:
            # Get date filters and class name
            start_date, end_date, period_name = self.get_date_filters(period)
            class_name = self.get_class_name(class_id, db)
            
            # Get student analytics data for CSV
            students_query = db.query(Student).filter(Student.is_active == True)
            if class_id:
                students_query = students_query.filter(Student.class_id == class_id)
            
            csv_data = []
            for student in students_query.all():
                student_records = db.query(AttendanceRecord).join(AttendanceSession).filter(
                    AttendanceRecord.student_id == student.id,
                    AttendanceSession.created_at >= start_date,
                    AttendanceSession.created_at <= end_date
                )
                if class_id:
                    student_records = student_records.filter(AttendanceSession.class_id == class_id)
                
                total_sessions = student_records.count()
                present_sessions = student_records.filter(AttendanceRecord.is_present == True).count()
                
                if total_sessions > 0:
                    attendance_rate = round((present_sessions / total_sessions) * 100, 1)
                    csv_data.append({
                        'Student_Name': student.name,
                        'Roll_Number': student.roll_no,
                        'PRN': student.prn,
                        'Total_Sessions': total_sessions,
                        'Present_Sessions': present_sessions,
                        'Absent_Sessions': total_sessions - present_sessions,
                        'Attendance_Percentage': attendance_rate
                    })
            
            # Create CSV content
            output = io.StringIO()
            if csv_data:
                fieldnames = csv_data[0].keys()
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(csv_data)
            
            # Create response
            filename = f"Attendance_{period}_{class_name}_{end_date.strftime('%Y%m%d')}.csv"
            
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8')),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        except Exception as e:
            logger.error(f"CSV export error: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")


# Create a global instance for easy import
attendance_exporter = AttendanceExporter()
